#!/usr/bin/env python3

import os
import zipfile
import re
import difflib
import tarfile
from openpyxl import load_workbook
from shutil import copy2 as copy_file

from formatting import *

projects_dir = 'Intro. a la prog. imperative-TP3-439361/'
dir_to_grade = 'To grade/'
dir_alr_graded = 'Already graded/'
marks = 'Notes_TP.xlsx'
mark_pattern = "/\*\s*Mark (\d+[\.\,]{0,1}\d*)\s*\*/"
num_of_students = 204
num_of_exercises = 10

marks_wb = load_workbook(marks)
tp_ws = marks_wb["TP3"]
# students = {}
# for k, row in enumerate(tp_ws.iter_rows(min_row=2, max_col=11, max_row=num_of_students+1)):
#    students[row[0].value] = [k+1] + list([el for el in row[1:]])
# ws['A1'] = 42

studentids = [cell[0].value for cell in marks_wb["All"]['A2':'A'+str(num_of_students+1)]]        
lastnames = [cell[0].value.lower() for cell in marks_wb["All"]['B2':'B'+str(num_of_students+1)]]
firstnames = [cell[0].value.lower() for cell in marks_wb["All"]['C2':'C'+str(num_of_students+1)]]

def get_name(studentnum):
    idx = studentids.index(studentnum)
    return (firstnames[idx] + ' ' + lastnames[idx]).title()

exs_to_grade_dirs, exs_alr_graded_dirs = ensure_dirs(dir_to_grade, dir_alr_graded, tp_ws, num_of_exercises)

basenames = ['exo', 'Exo', 'EXO', 'ex', 'Ex', 'EX']
numbers = [str(num) for num in range(1, num_of_exercises+1)]
suffixes = ['', 'bis', 'Bis', 'BIS', 'test']
extensions = ['', '.o', '.c~']
noproblem_files = ['{}{}{}{}'.format(basename, number, suffix, extension) for basename in basenames for number in numbers for suffix in suffixes for extension in extensions]
noproblem_files += ['#' + ok_file + '#' for ok_file in noproblem_files] + ['.DS_Store']


def match_folder_name_to_num(foldername):
    to_match = foldername[:foldername.find('_')].lower()
    try:
        to_match_by_surname = to_match[to_match.find(" ")+1:]
    except:
        to_match_by_surname = to_match
    matches_by_surname = difflib.get_close_matches(to_match_by_surname, lastnames)
    if len(matches_by_surname) == 0:
        for lastname in lastnames:
            if lastname in to_match_by_surname:
                matches_by_surname.append(lastname)
        sorted(matches_by_surname, key=lambda x: similar(to_match_by_surname, x), reverse=True)
    candidate_ids_by_surname_list = list()
    for m in matches_by_surname:
        for i, lastname in enumerate(lastnames):
            if m == lastname:
                candidate_ids_by_surname_list.append(studentids[i])
    if len(candidate_ids_by_surname_list) == 0:
        print_problem("The student with the folder {} hasn't been found.".format(foldername))
        return None
    elif len(candidate_ids_by_surname_list) == 1 or firstnames[studentids.index(candidate_ids_by_surname_list[0])] in to_match:
        return candidate_ids_by_surname_list[0]
    candidate_ids_by_surname = set(candidate_ids_by_surname_list)
    matches_by_firstname = difflib.get_close_matches(to_match, firstnames)
    candidate_ids_by_name = set([studentids[i] for i, name in enumerate(firstnames) if name in matches_by_firstname])
    candidate_ids = list(candidate_ids_by_surname.intersection(candidate_ids_by_name))
    if len(candidate_ids) == 1:
        return candidate_ids[0]
    elif candidate_ids:
        print_problem("The student with the folder {} cannot be uniquely identified. The potential candidates: {}.".format(foldername, "; ".join(candidate_ids)))
        return candidate_ids
    else:
        print_problem("The student with the folder {} hasn't been found.".format(foldername))
        return None


def not_graded_yet(student_id, exercise_idx):
    # exercise_idx is supposed to start at 1 and end at num_of_exercises
    if tp_ws[colnum_string(exercise_idx + 1) + str(studentids.index(student_id) + 2)].value != None:
        return False # It's already done
    return True # You still have to grade it 


def process_grades(dir_to_grade, dir_alr_graded):
    for ex_dir in os.listdir(dir_to_grade):
        if os.path.isdir(os.path.join(dir_to_grade, ex_dir)):
            ex_idx = exs_to_grade_dirs.index(os.path.join(dir_to_grade, ex_dir))
            col = colnum_string(ex_idx + 2)
            for subm_f in os.listdir(os.path.join(dir_to_grade, ex_dir)):
                if subm_f != '.DS_Store':
                    subm_f_loc = os.path.join(dir_to_grade, ex_dir, subm_f)
                    student_id = subm_f[:subm_f.find('_')]
                    row = str(studentids.index(int(student_id)) + 2)
                    with open(subm_f_loc, encoding='utf-8', errors='ignore') as f_to_grade:
                        f_contents = f_to_grade.read()
                        mark_match = re.search(mark_pattern, f_contents)
                        if mark_match != None:
                            marks_wb["TP3"][col + row] = float(mark_match.group(1))
                            marks_wb.save(marks)
                            os.rename(subm_f_loc, os.path.join(exs_alr_graded_dirs[ex_idx], subm_f))
                    
                    
def process_new_submissions(submissions_dir):
    verifystudentids = str()
    verifyfiles = str()
    for student in os.listdir(submissions_dir):
        submitted = ['-']*num_of_exercises
        problematic_files = list()
        if os.path.isdir(os.path.join(submissions_dir, student)):
            student_id = match_folder_name_to_num(student)
            if student_id is None or not student.startswith(get_name(student_id)):
                verifystudentids += "Please verify whether the student {} (folder {}) has the student number {}.\n".format(get_name(student_id), student, student_id)
            for submission in os.listdir(os.path.join(submissions_dir, student)):
                if submission.endswith('.zip') or submission.endswith('.tar.gz') or submission.endswith('.tar'):
                    where_to_unzip = os.path.join(submissions_dir, student, 'tmp')
                    archive = os.path.join(submissions_dir, student, submission)
                    if submission.endswith('.zip'):
                        arch_ref = zipfile.ZipFile(archive)
                    else:
                        arch_ref = tarfile.open(archive, 'r:gz' if submission.endswith('.tar.gz') else 'r:')
                    arch_ref.extractall(where_to_unzip)                   
                    for root, _, files in os.walk(where_to_unzip, topdown=False):
                        for f in files:
                            if f.endswith('.c') or f.endswith('.C'):
                                subm_file = os.path.join(root, f)
                                to_grade_file = os.path.join('{}', '{}_'.format(student_id) + f)
                                try:
                                    exercise_num = re.findall("(\d+)",f)[-1]
                                    submitted[int(exercise_num)-1] = '+'
                                    if not_graded_yet(student_id, int(exercise_num)):
                                        copy_file(subm_file, to_grade_file.format(exs_to_grade_dirs[int(exercise_num)-1]))
                                except:
                                    exercise_num = 'UNK'
                                    # print("Exercise {} (exercise #{}), has been found for student {} ({}), in the folder {}.".format(f, exercise_num, student_id, get_name(student_id), student))
                            elif f not in noproblem_files or exercise_num == 'UNK':
                                problematic_files.append(f)
            if problematic_files:
                verifyfiles += "File(s) {} have been found for student {} ({}), in the folder {}.\n".format(", ".join(problematic_files), student_id, get_name(student_id), student)
            print_status("Student {} ({}) submitted the following exercises: {} (the folder {}).".format(student_id, get_name(student_id), " ".join(submitted), student))
            for ex_idx, status in enumerate(submitted):
                if status == '-':
                    col = colnum_string(ex_idx + 2)
                    row = str(studentids.index(int(student_id)) + 2)
                    marks_wb["TP3"][col + row] = 0
                    marks_wb.save(marks)
    print_problem(verifystudentids)
    print_problem(verifyfiles)

process_new_submissions(projects_dir)
process_grades(dir_to_grade, dir_alr_graded)